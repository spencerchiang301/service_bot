"""
範本產生器：依格式回傳 bytes + (filename, media_type)
"""
import csv
import io

# ── 範例資料 ──────────────────────────────────────────────────────────────────

SAMPLE_PRODUCTS = [
    ("美式咖啡", "大杯", 80, ""),
    ("美式咖啡", "中杯", 65, ""),
    ("拿鐵",     "大杯", 110, "可選燕麥奶，加價 15 元"),
    ("起司蛋糕", "",    120, "每片"),
    ("可頌麵包", "",     65, "每日現烤，售完為止"),
]

SAMPLE_SERVICES = [
    ("外送服務", "訂單滿 300 元免運費，未滿收 50 元", "50（未滿額）", "外送範圍：本店 3 公里內"),
    ("外送時間", "每日 11:00–21:00 提供外送", "", ""),
    ("預訂服務", "可提前一天預訂，最少 10 人份", "", "請來電或加 LINE"),
    ("客製化蛋糕", "可選口味：原味、巧克力、抹茶、草莓", "4吋580/6吋880/8吋1280", "需提前 3 天預訂"),
]

SAMPLE_INFO = [
    ("營業時間（平日）", "週一至週五 09:00–21:00"),
    ("營業時間（假日）", "週六、週日 10:00–22:00"),
    ("地址",             "台北市中山區範例路 1 號"),
    ("電話",             "02-1234-5678"),
    ("LINE 官方帳號",    "@myshop"),
    ("Email",            "hello@myshop.com"),
]


# ── Excel ─────────────────────────────────────────────────────────────────────

def gen_xlsx() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    BLUE   = "2563EB"
    GRAY   = "64748B"
    TEAL   = "0F766E"
    SAMPLE = "94A3B8"

    def hdr(cell, color):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        s = Side(style="thin", color="E2E8F0")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    def data(cell, sample=True):
        cell.font = Font(color=SAMPLE if sample else "1E293B", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        s = Side(style="thin", color="E2E8F0")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 說明頁 ──
    ws0 = wb.create_sheet("📋 說明")
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 80
    lines = [
        ("客服機器人資料填寫說明", True),
        ("", False),
        ("本檔案共有 3 個工作表，請依照各分頁說明填入您的資料：", False),
        ("", False),
        ("🛍️ 商品   — 填寫商品名稱、規格、價格", False),
        ("🔧 服務   — 填寫服務項目與費用說明", False),
        ("ℹ️  基本資訊 — 填寫營業時間、地址、聯絡方式", False),
        ("", False),
        ("注意事項：", True),
        ("• 灰色文字為範例，直接覆蓋成您的資料即可", False),
        ("• 價格欄位只填數字，不需加 $ 符號", False),
        ("• 備註欄位可空白", False),
        ("• 填完後回到管理後台上傳，再按「重新建立索引」", False),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws0.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=12 if bold else 11, color="2563EB" if bold else "1E293B")
        c.alignment = Alignment(wrap_text=True)
        ws0.row_dimensions[i].height = 22 if txt else 8

    # ── 商品 ──
    ws1 = wb.create_sheet("🛍️ 商品")
    ws1.sheet_view.showGridLines = False
    hdrs = ["商品名稱 *", "規格／尺寸", "價格（元）*", "備註"]
    widths = [25, 18, 16, 35]
    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        hdr(ws1.cell(1, c, h), BLUE)
        ws1.column_dimensions[ws1.cell(1, c).column_letter].width = w
    ws1.row_dimensions[1].height = 28
    for r, row in enumerate(SAMPLE_PRODUCTS, 2):
        for c, v in enumerate(row, 1):
            data(ws1.cell(r, c, v))
    ws1.freeze_panes = "A2"

    # ── 服務 ──
    ws2 = wb.create_sheet("🔧 服務")
    ws2.sheet_view.showGridLines = False
    hdrs2 = ["服務名稱 *", "服務說明 *", "費用", "備註"]
    widths2 = [20, 45, 18, 30]
    for c, (h, w) in enumerate(zip(hdrs2, widths2), 1):
        hdr(ws2.cell(1, c, h), GRAY)
        ws2.column_dimensions[ws2.cell(1, c).column_letter].width = w
    ws2.row_dimensions[1].height = 28
    for r, row in enumerate(SAMPLE_SERVICES, 2):
        for c, v in enumerate(row, 1):
            data(ws2.cell(r, c, v))
    ws2.freeze_panes = "A2"

    # ── 基本資訊 ──
    ws3 = wb.create_sheet("ℹ️ 基本資訊")
    ws3.sheet_view.showGridLines = False
    for c, (h, w) in enumerate(zip(["項目 *", "內容 *"], [22, 55]), 1):
        hdr(ws3.cell(1, c, h), TEAL)
        ws3.column_dimensions[ws3.cell(1, c).column_letter].width = w
    ws3.row_dimensions[1].height = 28
    for r, row in enumerate(SAMPLE_INFO, 2):
        for c, v in enumerate(row, 1):
            data(ws3.cell(r, c, v))
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CSV ───────────────────────────────────────────────────────────────────────

def gen_csv() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["# 使用說明：將此檔案填寫完後，上傳至管理後台的知識庫"])
    w.writerow(["# 類型欄位：商品 / 服務 / 基本資訊"])
    w.writerow([])
    w.writerow(["類型", "名稱", "規格／尺寸", "價格（元）", "說明", "備註"])

    for row in SAMPLE_PRODUCTS:
        w.writerow(["商品", row[0], row[1], row[2], "", row[3]])

    w.writerow([])
    for row in SAMPLE_SERVICES:
        w.writerow(["服務", row[0], "", "", row[1], f"費用：{row[2]}　{row[3]}".strip("　")])

    w.writerow([])
    for row in SAMPLE_INFO:
        w.writerow(["基本資訊", row[0], "", "", row[1], ""])

    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


# ── TXT ───────────────────────────────────────────────────────────────────────

def gen_txt() -> bytes:
    lines = [
        "================================================================",
        "  客服機器人資料範本",
        "  填寫說明：刪除範例內容，填入您自己的資料後儲存並上傳",
        "================================================================",
        "",
        "【商品價格表】",
        "格式：商品名稱（規格）：價格元　備註",
        "------------------------------------------------------------",
    ]
    for name, spec, price, note in SAMPLE_PRODUCTS:
        label = f"{name}（{spec}）" if spec else name
        line = f"{label}：{price} 元"
        if note:
            line += f"　{note}"
        lines.append(line)

    lines += [
        "",
        "【服務項目】",
        "格式：服務名稱：說明　費用　備註",
        "------------------------------------------------------------",
    ]
    for name, desc, fee, note in SAMPLE_SERVICES:
        line = f"{name}：{desc}"
        if fee:
            line += f"　費用：{fee}"
        if note:
            line += f"　{note}"
        lines.append(line)

    lines += [
        "",
        "【基本資訊】",
        "格式：項目：內容",
        "------------------------------------------------------------",
    ]
    for item, content in SAMPLE_INFO:
        lines.append(f"{item}：{content}")

    lines += ["", "================================================================"]
    return "\n".join(lines).encode("utf-8")


# ── PDF ───────────────────────────────────────────────────────────────────────

def gen_pdf() -> bytes:
    import platform
    from fpdf import FPDF

    # 找系統中文字型
    FONT_PATHS = [
        "/System/Library/Fonts/STHeiti Medium.ttc",          # macOS
        "/System/Library/Fonts/PingFang.ttc",                # macOS newer
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",     # Linux
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "C:/Windows/Fonts/msjh.ttc",                         # Windows
    ]
    font_path = next((p for p in FONT_PATHS if __import__('os').path.exists(p)), None)

    class PDF(FPDF):
        def __init__(self, fp):
            super().__init__(orientation="L", format="A4")
            self.add_font("CJK", fname=fp)
            self.set_font("CJK", size=10)

        def header(self):
            self.set_font("CJK", size=15)
            self.set_text_color(37, 99, 235)
            self.cell(0, 10, "客服機器人  資料填寫範本", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("CJK", size=9)
            self.set_text_color(100, 116, 139)
            self.cell(0, 6, "填入您的商品與服務資料後，上傳至管理後台 → 知識庫 → 重新建立索引", align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(4)

        def section(self, title, color):
            self.set_fill_color(*color)
            self.set_text_color(255, 255, 255)
            self.set_font("CJK", size=11)
            self.cell(0, 8, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
            self.set_text_color(30, 41, 59)
            self.ln(2)

        def table_header(self, cols, widths):
            self.set_fill_color(241, 245, 249)
            self.set_text_color(100, 116, 139)
            self.set_font("CJK", size=9)
            for col, w in zip(cols, widths):
                self.cell(w, 7, col, border=1, fill=True)
            self.ln()

        def table_row(self, values, widths, shade=False):
            if shade:
                self.set_fill_color(248, 250, 252)
            else:
                self.set_fill_color(255, 255, 255)
            self.set_text_color(30, 41, 59)
            self.set_font("CJK", size=9)
            for v, w in zip(values, widths):
                self.cell(w, 7, str(v), border=1, fill=True)
            self.ln()

    if not font_path:
        # fallback：純英文版本
        pdf_fb = FPDF(orientation="L", format="A4")
        pdf_fb.add_page()
        pdf_fb.set_font("Helvetica", size=11)
        pdf_fb.cell(0, 10, "Service Bot Data Template", new_x="LMARGIN", new_y="NEXT")
        pdf_fb.set_font("Helvetica", size=9)
        for name, spec, price, note in SAMPLE_PRODUCTS:
            pdf_fb.cell(0, 7, f"Product: {name} {spec}  Price: {price}  Note: {note}", new_x="LMARGIN", new_y="NEXT")
        return bytes(pdf_fb.output())

    pdf = PDF(font_path)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_margins(15, 15, 15)

    pdf.section("商品價格表", (37, 99, 235))
    pdf.table_header(["商品名稱", "規格／尺寸", "價格（元）", "備註"], [70, 40, 35, 115])
    for i, (name, spec, price, note) in enumerate(SAMPLE_PRODUCTS):
        pdf.table_row([name, spec, price, note], [70, 40, 35, 115], shade=i % 2 == 1)

    pdf.ln(5)
    pdf.section("服務項目", (100, 116, 139))
    pdf.table_header(["服務名稱", "服務說明", "費用", "備註"], [45, 110, 35, 70])
    for i, (name, desc, fee, note) in enumerate(SAMPLE_SERVICES):
        pdf.table_row([name, desc, fee, note], [45, 110, 35, 70], shade=i % 2 == 1)

    pdf.ln(5)
    pdf.section("基本資訊", (15, 118, 110))
    pdf.table_header(["項目", "內容"], [60, 200])
    for i, (item, content) in enumerate(SAMPLE_INFO):
        pdf.table_row([item, content], [60, 200], shade=i % 2 == 1)

    return bytes(pdf.output())


# ── Dispatcher ────────────────────────────────────────────────────────────────

FORMATS = {
    "xlsx": (gen_xlsx, "shop_template.xlsx",
             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "csv":  (gen_csv,  "shop_template.csv",  "text/csv; charset=utf-8"),
    "txt":  (gen_txt,  "shop_template.txt",  "text/plain; charset=utf-8"),
    "pdf":  (gen_pdf,  "shop_template.pdf",  "application/pdf"),
}


def generate(fmt: str) -> tuple[bytes, str, str]:
    """回傳 (data, filename, media_type)"""
    if fmt not in FORMATS:
        raise ValueError(f"Unknown format: {fmt}")
    fn, filename, media_type = FORMATS[fmt]
    return fn(), filename, media_type
