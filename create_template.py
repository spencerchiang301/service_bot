"""
執行此腳本產生 Excel 範本：
    python create_template.py
會在 data/ 資料夾產生 shop_data.xlsx
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── 顏色 ──────────────────────────────────────────────
BLUE_HDR   = "2563EB"
GRAY_HDR   = "64748B"
LIGHT_BLUE = "EFF6FF"
LIGHT_GRAY = "F8FAFC"
SAMPLE_FONT_COLOR = "94A3B8"  # 範例資料用淺灰，方便辨識

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=BLUE_HDR):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = hdr_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def style_cell(cell, bg="FFFFFF", sample=False):
    cell.font = Font(color=SAMPLE_FONT_COLOR if sample else "1E293B", size=10)
    cell.fill = hdr_fill(bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border()


# ── Sheet 1：說明 ─────────────────────────────────────
def make_instructions(ws):
    ws.title = "📋 說明"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 80

    lines = [
        ("客服機器人資料填寫說明", True),
        ("", False),
        ("本檔案共有 4 個工作表，請依照各分頁說明填寫：", False),
        ("", False),
        ("🛍️ 商品   — 填寫所有販售的商品名稱、規格與價格", False),
        ("🔧 服務   — 填寫提供的服務項目（外送、預訂、客製等）", False),
        ("ℹ️  基本資訊 — 填寫營業時間、地址、電話等店家資訊", False),
        ("", False),
        ("填寫注意事項：", True),
        ("• 灰色文字為範例，請直接覆蓋或刪除後填入自己的資料", False),
        ("• 價格欄位只需填數字，不需加 $ 符號", False),
        ("• 備註欄位可空白", False),
        ("• 完成後執行：python ingest.py --reset  即可更新機器人的知識庫", False),
    ]

    for i, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if bold else 11,
                         color="1E293B" if not bold else "2563EB")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[i].height = 22 if text else 10


# ── Sheet 2：商品 ─────────────────────────────────────
def make_products(ws):
    ws.title = "🛍️ 商品"
    ws.sheet_view.showGridLines = False

    headers = ["商品名稱 *", "規格／尺寸", "價格（元）*", "備註"]
    col_widths = [25, 18, 16, 35]
    samples = [
        ("美式咖啡", "大杯", 80, ""),
        ("美式咖啡", "中杯", 65, ""),
        ("拿鐵", "大杯", 110, "可選燕麥奶，加價 15 元"),
        ("起司蛋糕", "", 120, "每片"),
        ("可頌麵包", "", 65, "每日現烤，售完為止"),
    ]

    # 標題列
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    # 範例資料
    for r, row in enumerate(samples, 2):
        bg = LIGHT_BLUE if r % 2 == 0 else "FFFFFF"
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_cell(cell, bg=bg, sample=True)
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"


# ── Sheet 3：服務 ─────────────────────────────────────
def make_services(ws):
    ws.title = "🔧 服務"
    ws.sheet_view.showGridLines = False

    headers = ["服務名稱 *", "服務說明 *", "費用", "備註"]
    col_widths = [20, 45, 18, 30]
    samples = [
        ("外送服務", "訂單滿 300 元免運費，未滿收 50 元運費", "50（未滿額）", "外送範圍：本店 3 公里內"),
        ("外送時間", "每日 11:00 – 21:00 提供外送", "", ""),
        ("預訂服務", "可提前一天預訂，最少 10 人份起", "", "請來電或加 LINE"),
        ("客製化蛋糕", "可選口味：原味、巧克力、抹茶、草莓", "4吋580 / 6吋880 / 8吋1280", "需提前 3 天預訂"),
    ]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell, bg=GRAY_HDR)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    for r, row in enumerate(samples, 2):
        bg = LIGHT_GRAY if r % 2 == 0 else "FFFFFF"
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_cell(cell, bg=bg, sample=True)
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "A2"


# ── Sheet 4：基本資訊 ─────────────────────────────────
def make_info(ws):
    ws.title = "ℹ️ 基本資訊"
    ws.sheet_view.showGridLines = False

    headers = ["項目 *", "內容 *"]
    col_widths = [22, 55]
    samples = [
        ("營業時間（平日）", "週一至週五 09:00 – 21:00"),
        ("營業時間（假日）", "週六、週日 10:00 – 22:00"),
        ("國定假日", "照常營業"),
        ("地址", "台北市中山區範例路 1 號"),
        ("電話", "02-1234-5678"),
        ("LINE 官方帳號", "@myshop"),
        ("Email", "hello@myshop.com"),
        ("停車", "店門口有 3 格停車位，免費使用"),
    ]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell, bg="0F766E")  # teal
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    for r, row in enumerate(samples, 2):
        bg = "F0FDFA" if r % 2 == 0 else "FFFFFF"
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_cell(cell, bg=bg, sample=True)
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"


# ── 主程式 ────────────────────────────────────────────
if __name__ == "__main__":
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 刪除預設空白 sheet

    make_instructions(wb.create_sheet())
    make_products(wb.create_sheet())
    make_services(wb.create_sheet())
    make_info(wb.create_sheet())

    out = Path("data") / "shop_data.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"✅ 範本已產生：{out}")
    print("   請用 Excel 開啟並填入您的資料，完成後執行：")
    print("   python ingest.py --reset")
